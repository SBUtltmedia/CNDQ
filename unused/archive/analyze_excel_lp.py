import openpyxl
import sys

# Load the Excel workbook
wb = openpyxl.load_workbook('Deicer and Solvent Model Blank Template 19Dect2025.xlsm', data_only=False)

print("=== Excel Workbook Analysis ===\n")
print(f"Worksheets: {wb.sheetnames}\n")

# Check each sheet for LP-related content
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")

    # Look for key terms in the sheet
    key_terms = ['shadow', 'dual', 'price', 'constraint', 'objective', 'solver', 'optimal']

    found_cells = []
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=20):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_lower = cell.value.lower()
                for term in key_terms:
                    if term in cell_lower:
                        found_cells.append({
                            'cell': cell.coordinate,
                            'value': cell.value,
                            'term': term
                        })
                        break

    if found_cells:
        print("Found LP-related cells:")
        for item in found_cells:
            print(f"  {item['cell']}: {item['value']}")

    # Print first 30 rows to see structure
    print(f"\nFirst 30 rows of {sheet_name}:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10), 1):
        row_data = []
        for cell in row:
            val = cell.value
            if val is not None:
                # Show formulas if available
                if hasattr(cell, 'value') and isinstance(val, str) and val.startswith('='):
                    row_data.append(f"{cell.coordinate}={val}")
                else:
                    row_data.append(f"{cell.coordinate}:{val}")
        if row_data:
            print(f"  Row {row_idx}: {', '.join(row_data)}")

print("\n=== Defined Names (Named Ranges) ===")
if wb.defined_names:
    for name in wb.defined_names:
        print(f"  {name.name}: {name.value}")
