#!/usr/bin/env python3
"""
Extract information from the CNDQ Excel template file.
This script extracts:
- Sheet names and structures
- Cell formulas and values
- VBA macro code
- Report templates
"""

import openpyxl
from openpyxl import load_workbook
import json
import sys
import os
from pathlib import Path

def extract_sheet_info(wb):
    """Extract basic information about all sheets."""
    sheets_info = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column

        # Sample first few rows to understand structure
        sample_data = []
        for row in range(1, min(6, max_row + 1)):
            row_data = []
            for col in range(1, min(10, max_col + 1)):
                cell = ws.cell(row, col)
                row_data.append({
                    'value': cell.value,
                    'has_formula': cell.data_type == 'f'
                })
            sample_data.append(row_data)

        sheets_info[sheet_name] = {
            'max_row': max_row,
            'max_col': max_col,
            'sample_data': sample_data
        }

    return sheets_info

def extract_team_sheet_structure(wb, sheet_name='Team 1'):
    """Extract the structure of a team sheet to understand the template."""
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    structure = {
        'objective_section': {},
        'constraints': [],
        'variables': {},
        'sales_section': {},
        'purchases_section': {},
        'financial_summary': {}
    }

    # Extract key cells and their formulas
    for row in range(1, min(50, ws.max_row + 1)):
        for col in range(1, min(50, ws.max_column + 1)):
            cell = ws.cell(row, col)
            if cell.value and (isinstance(cell.value, str) or cell.data_type == 'f'):
                coord = f"{cell.column_letter}{cell.row}"

                # Try to categorize
                if row <= 4:
                    structure['objective_section'][coord] = {
                        'value': cell.value,
                        'formula': cell.value if cell.data_type == 'f' else None
                    }

    return structure

def extract_scoreboard_structure(wb):
    """Extract the scoreboard structure."""
    if 'Scoreboard' not in wb.sheetnames:
        return None

    ws = wb['Scoreboard']
    scoreboard = {
        'headers': [],
        'team_formulas': []
    }

    # Get headers (likely row 1)
    for col in range(1, min(10, ws.max_column + 1)):
        cell = ws.cell(1, col)
        if cell.value:
            scoreboard['headers'].append({
                'column': cell.column_letter,
                'value': cell.value
            })

    # Get first few team rows to see formula pattern
    for row in range(2, min(7, ws.max_row + 1)):
        team_row = {}
        for col in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row, col)
            coord = f"{cell.column_letter}{cell.row}"
            team_row[coord] = {
                'value': cell.value,
                'formula': cell.value if cell.data_type == 'f' else None
            }
        scoreboard['team_formulas'].append(team_row)

    return scoreboard

def extract_vba_macros(excel_path):
    """Extract VBA macro code from the Excel file."""
    try:
        # xlsm files are actually ZIP files
        import zipfile
        import xml.etree.ElementTree as ET

        macros = {}

        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            # List all files in the archive
            file_list = zip_ref.namelist()

            # Look for VBA project files
            vba_files = [f for f in file_list if 'vba' in f.lower() or 'xl/vbaProject.bin' in f]

            if vba_files:
                macros['vba_files_found'] = vba_files
                macros['note'] = 'VBA binary found. Use olevba tool to extract: pip install oletools && olevba "file.xlsm"'

            # Check for worksheet XML that might have button definitions
            xl_files = [f for f in file_list if f.startswith('xl/')]
            macros['xl_structure'] = xl_files[:20]  # First 20 files

        return macros
    except Exception as e:
        return {'error': str(e)}

def extract_report_templates(wb):
    """Extract Answer Report and Sensitivity Report structures."""
    reports = {}

    # Answer Report
    if 'Answer Report 1' in wb.sheetnames:
        ws = wb['Answer Report 1']
        answer_report = []
        for row in range(1, min(30, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(10, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value:
                    row_data.append(str(cell.value))
            if row_data:
                answer_report.append(row_data)
        reports['answer_report'] = answer_report

    # Sensitivity Report
    if 'Sensitivity Report 1' in wb.sheetnames:
        ws = wb['Sensitivity Report 1']
        sensitivity_report = []
        for row in range(1, min(30, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(10, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value:
                    row_data.append(str(cell.value))
            if row_data:
                sensitivity_report.append(row_data)
        reports['sensitivity_report'] = sensitivity_report

    return reports

def main():
    # Path to Excel file
    excel_path = Path(__file__).parent.parent / "unused" / "archive" / "Deicer and Solvent Model Blank Template 19Dect2025.xlsm"

    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)

    print(f"Loading Excel file: {excel_path}")
    print("=" * 80)

    # Load workbook (data_only=False to see formulas)
    wb = load_workbook(excel_path, data_only=False)

    print("\n1. SHEET NAMES:")
    print("-" * 80)
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"{i:2d}. {name}")

    print("\n2. SHEETS INFO:")
    print("-" * 80)
    sheets_info = extract_sheet_info(wb)
    for name, info in sheets_info.items():
        print(f"\n{name}:")
        print(f"  Dimensions: {info['max_row']} rows x {info['max_col']} columns")

    print("\n3. TEAM SHEET STRUCTURE (Team 1 as example):")
    print("-" * 80)
    team_structure = extract_team_sheet_structure(wb, 'Team 1')
    if team_structure:
        print(json.dumps(team_structure, indent=2, default=str))

    print("\n4. SCOREBOARD STRUCTURE:")
    print("-" * 80)
    scoreboard = extract_scoreboard_structure(wb)
    if scoreboard:
        print(json.dumps(scoreboard, indent=2, default=str))

    print("\n5. REPORT TEMPLATES:")
    print("-" * 80)
    reports = extract_report_templates(wb)
    for report_name, content in reports.items():
        print(f"\n{report_name}:")
        for row in content[:15]:  # First 15 rows
            print("  " + " | ".join(row))

    print("\n6. VBA MACROS:")
    print("-" * 80)
    macros = extract_vba_macros(excel_path)
    print(json.dumps(macros, indent=2))

    # Save detailed output to JSON file
    output_path = Path(__file__).parent / "excel_extraction.json"
    output_data = {
        'sheets': list(wb.sheetnames),
        'sheets_info': sheets_info,
        'team_structure': team_structure,
        'scoreboard': scoreboard,
        'reports': reports,
        'vba_info': macros
    }

    with open(output_path, 'w') as f:
        json.dump(output_data, f, indent=2, default=str)

    print(f"\n\nDetailed output saved to: {output_path}")
    print("=" * 80)

if __name__ == '__main__':
    main()
